from urllib import request

from django.shortcuts import render, redirect


from rest_framework import viewsets
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter

from .models import Badge,ColumnMapping
from .serializers import BadgeSerializer,InvitationSerializer,Invitation

class BadgeViewSet(viewsets.ModelViewSet):

    serializer_class = BadgeSerializer

    filter_backends = [
        DjangoFilterBackend,
        SearchFilter
    ]

    filterset_fields = [
        "status",
        "ticket"
    ]

    search_fields = [
        "first_name",
        "last_name",
        "email",
        "company_name"
    ]

    def get_queryset(self):
        user = self.request.user
        if hasattr(user, 'exhibitor'):
            return Badge.objects.filter(
                exhibitor=user.exhibitor
            ).order_by("-id")
        return Badge.objects.all().order_by("-id")


from rest_framework.views import APIView
from rest_framework.response import Response
from django.contrib.auth import login
from django.db.models import Sum



class DashboardAPIView(APIView):

    def get(self, request):

        exhibitor = request.user.exhibitor

        allocated = exhibitor.allocated_badges

        invited = Invitation.objects.filter(
            exhibitor=exhibitor
        ).count()

        confirmed_invitation = Invitation.objects.filter(
            exhibitor=exhibitor,
            status="Confirmed"
        ).count()
        

        
        confirmed_upload = UploadRecord.objects.filter(
            batch__exhibitor=exhibitor,
            is_valid=True
            
        ).count()

        confirmed = confirmed_invitation + confirmed_upload
        

        available = max(
            allocated - confirmed,
            0
        )

        ticket_summary = []

        tickets = TicketType.objects.all()

        for ticket in tickets:

            invited_count = Invitation.objects.filter(
                exhibitor=exhibitor,
                ticket=ticket
            ).count()

            used = Invitation.objects.filter(
                exhibitor=exhibitor,
                ticket=ticket,
                status="Confirmed"
            ).count()

            ticket_summary.append({
                "ticket_name": ticket.name,
                "allocated": allocated,
                "invited": invited_count,
                "used": used,
                "available": max(allocated - used, 0)
            })

        return Response({

            "allocated_badges": allocated,

            "total_invited_count": invited,

            "confirmed_count": confirmed,

            "available_badge_balance": available,

            "ticket_summary": ticket_summary

        })
        
from rest_framework import viewsets

class InvitationViewSet(viewsets.ModelViewSet):

    queryset = Invitation.objects.all().order_by("-id")
    serializer_class = InvitationSerializer
    
    
    
    
    
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework.response import Response

class RegistrationAPIView(APIView):

    def get(self, request, token):

        invitation = get_object_or_404(
            Invitation,
            invitation_token=token
        )

        return Response({
            "first_name": invitation.first_name,
            "last_name": invitation.last_name,
            "email": invitation.email,
            "ticket": invitation.ticket.id
        })
        
        
class RegistrationAPIView(APIView):

    def get(self, request, token):

        invitation = get_object_or_404(
            Invitation,
            invitation_token=token
        )

        return Response({
            "first_name": invitation.first_name,
            "last_name": invitation.last_name,
            "email": invitation.email,
            "ticket": invitation.ticket.id
        })

    def post(self, request, token):

        invitation = get_object_or_404(
            Invitation,
            invitation_token=token
        )

        if invitation.is_used:
            return Response(
                {"error": "Invitation already used"},
                status=400
            )

        badge = Badge.objects.create(
            first_name=invitation.first_name,
            last_name=invitation.last_name,
            email=invitation.email,
            ticket=invitation.ticket,

            job_title=request.data.get("job_title"),
            company_name=request.data.get("company_name"),
            phone_number=request.data.get("phone_number"),
            country_of_residence=request.data.get(
                "country_of_residence"
            ),
            nationality=request.data.get("nationality"),

            status="confirmed"
        )

        invitation.is_used = True
        invitation.save()

        return Response({
            "message": "Registration completed",
            "badge_id": badge.id
        })
        
        

from rest_framework.views import APIView
from rest_framework.response import Response
from .models import Badge

class BulkDeleteBadgeAPIView(APIView):

    def post(self, request):
        exhibitor = resolve_exhibitor_from_request(request)
        if not exhibitor:
            return Response({"error": "Not authenticated."}, status=401)

        ids = request.data.get("ids", [])

        deleted_count, _ = Badge.objects.filter(
             id__in=ids,
             exhibitor=exhibitor
        ).delete()

        return Response({
            "deleted_count": deleted_count
        })


# ─── Single UploadRecord Delete ──────────────────────────────────────────────

class DeleteUploadRecordAPIView(APIView):
    """DELETE /upload-record/<int:record_id>/delete/
    Deletes a single UploadRecord row belonging to the logged-in exhibitor.
    """

    def delete(self, request, record_id):
        from .models import UploadRecord

        exhibitor = resolve_exhibitor_from_request(request)

        try:
            record = UploadRecord.objects.select_related("batch").get(
                id=record_id
            )
        except UploadRecord.DoesNotExist:
            return Response({"error": "Record not found."}, status=404)

        # Security: only the owning exhibitor may delete
        if exhibitor and record.batch.exhibitor != exhibitor:
            return Response({"error": "Not authorised."}, status=403)

        record.delete()
        return Response({"deleted": True, "record_id": record_id})


# ─── Delete SELECTED UploadRecords (checked rows only) ───────────────────────

class DeleteSelectedUploadRecordsAPIView(APIView):
    """POST /upload-records/delete-selected/
    Deletes only the UploadRecord rows whose IDs are passed in {"ids": [1,2,3]}.
    Scoped to the logged-in exhibitor for security.
    """

    def post(self, request):
        from .models import UploadRecord

        exhibitor = resolve_exhibitor_from_request(request)

        if not exhibitor:
            return Response({"error": "Not authenticated."}, status=401)

        ids = request.data.get("ids", [])

        if not ids:
            return Response({"error": "No IDs provided."}, status=400)

        # Only delete records that actually belong to this exhibitor
        qs = UploadRecord.objects.filter(
            id__in=ids,
            batch__exhibitor=exhibitor
        )

        deleted_count, _ = qs.delete()

        return Response({"deleted_count": deleted_count})


# ─── Bulk Delete ALL UploadRecords for an Exhibitor ──────────────────────────

class BulkDeleteUploadRecordAPIView(APIView):
    """DELETE /upload-records/bulk-delete/
    Deletes EVERY UploadRecord that belongs to the logged-in exhibitor in
    chunks of 5 000 so it handles 3-lakh rows without memory issues.
    Optionally accepts {"batch_id": <int>} in the body to restrict deletion
    to a single upload batch.
    """

    def delete(self, request):
        from .models import UploadRecord, UploadBatch

        exhibitor = resolve_exhibitor_from_request(request)

        if not exhibitor:
            return Response({"error": "Not authenticated."}, status=401)

        batch_id = request.data.get("batch_id")
        CHUNK = 5000
        total_deleted = 0

        if batch_id:
            try:
                batch = UploadBatch.objects.get(id=batch_id, exhibitor=exhibitor)
            except UploadBatch.DoesNotExist:
                return Response({"error": "Batch not found."}, status=404)
            qs = UploadRecord.objects.filter(batch=batch)
        else:
            # Wipe everything for this exhibitor
            qs = UploadRecord.objects.filter(batch__exhibitor=exhibitor)
            # Wipes all Badge creations for this exhibitor
            Badge.objects.filter(exhibitor=exhibitor).delete()

        # Chunked delete – safe for 3-lakh+ rows
        while True:
            id_chunk = list(qs.values_list("id", flat=True)[:CHUNK])
            if not id_chunk:
                break
            deleted, _ = UploadRecord.objects.filter(id__in=id_chunk).delete()
            total_deleted += deleted

        return Response({
            "deleted_count": total_deleted,
            "batch_id": batch_id
        })


from openpyxl import Workbook
from django.http import HttpResponse

class ExportBadgesAPIView(APIView):

    def get(self, request):

        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Badges"

        sheet.append([
            "Name",
            "Company Name",
            "Job Title",
            "Ticket Name",
            "Status",
            "Phone",
            "Email"
        ])

        exhibitor = getattr(request.user, 'exhibitor', None)
        if exhibitor:
            badges = Badge.objects.filter(exhibitor=exhibitor).select_related('ticket')
        else:
            badges = Badge.objects.all().select_related('ticket')

        for badge in badges:

            sheet.append([
                f"{badge.first_name} {badge.last_name}",
                badge.company_name,
                badge.job_title,
                badge.ticket.name if badge.ticket else '',
                badge.status,
                badge.phone_number,
                badge.email
            ])

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )

        response[
            "Content-Disposition"
        ] = 'attachment; filename="badges.xlsx"'

        workbook.save(response)

        return response
    
from openpyxl import load_workbook
from rest_framework.views import APIView
from rest_framework.response import Response
import re
import threading
from io import BytesIO

def resolve_exhibitor_from_request(request):
    print("=" * 60)
    print("request.user:", request.user)
    print("authenticated:", request.user.is_authenticated)

    if getattr(request.user, 'is_authenticated', False):
        print("Trying request.user.exhibitor...")
        exhibitor = getattr(request.user, 'exhibitor', None)
        print("Found exhibitor:", exhibitor)

        if exhibitor:
            return exhibitor

    username = request.session.get('username')
    print("Session username:", username)

    if username:
        try:
            from django.contrib.auth.models import User

            user = User.objects.get(username=username)
            print("Session user:", user)

            exhibitor = getattr(user, 'exhibitor', None)
            print("Session exhibitor:", exhibitor)

            if exhibitor:
                return exhibitor

        except Exception as e:
            print(e)

    return None

def process_upload_batch(batch_id, file_name, file_bytes, columns=None):
    batch = UploadBatch.objects.get(id=batch_id)

    try:
        is_csv = file_name.lower().endswith('.csv')
        rows_data = []

        if is_csv:
            import csv
            from io import StringIO
            try:
                text = file_bytes.decode('utf-8-sig')
            except UnicodeDecodeError:
                text = file_bytes.decode('latin-1')

            reader = csv.reader(StringIO(text))
            try:
                file_columns = next(reader)
            except StopIteration:
                file_columns = []
            
            # Clean columns: strip whitespace and ignore None/empty
            file_columns = [str(c).strip() if c is not None else "" for c in file_columns]

            if columns is None or not columns:
                columns = file_columns
            else:
                columns = [str(c).strip() if c is not None else "" for c in columns]

            for row in reader:
                rows_data.append(row)
        else:
            workbook = load_workbook(
                BytesIO(file_bytes),
                read_only=True,
                data_only=True
            )
            sheet = workbook.active

            if columns is None or not columns:
                columns = []
                for cell in sheet[1]:
                    val = str(cell.value).strip() if cell.value is not None else ""
                    columns.append(val)
            else:
                columns = [str(c).strip() if c is not None else "" for c in columns]

            for row in sheet.iter_rows(
                min_row=2,
                values_only=True
            ):
                rows_data.append(row)

            workbook.close()

        uploaded_emails = set()

        total_records = 0
        valid_records = 0
        invalid_records = 0
        records_to_create = []
        batch_size = 2000

        for row in rows_data:
            row_data = dict(zip(columns, row))

            errors = []

            # Clean and normalize fields to stripped strings or None
            def clean_field(val):
                if val is None:
                    return None
                s = str(val).strip()
                return s if s else None

            first_name = clean_field(row_data.get("First Name"))
            last_name = clean_field(row_data.get("Last Name"))
            email = clean_field(row_data.get("Email"))
            phone = clean_field(row_data.get("Phone"))
            job_title = clean_field(row_data.get("Job Title"))
            company = clean_field(row_data.get("Company"))

            if not first_name:
                errors.append("First Name is required")

            if not last_name:
                errors.append("Last Name is required")

            if not email:
                errors.append("Email is required")
            else:
                email_pattern = (
                    r'^[A-Za-z0-9._%+-]+'
                    r'@[A-Za-z0-9.-]+'
                    r'\.[A-Za-z]{2,}$'
                )

                if not re.match(email_pattern, email):
                    errors.append("Invalid email format")
                elif email in uploaded_emails:
                    errors.append("Duplicate email in uploaded file")

            if not phone:
                errors.append("Phone is required")
            elif not phone.isdigit():
                errors.append("Phone must contain only digits")
            elif len(phone) < 10:
                errors.append("Phone must be at least 10 digits")

            if not job_title:
                errors.append("Job Title is required")

            if not company:
                errors.append("Company is required")

            # Calculate validity first
            is_valid = len(errors) == 0

            # Keep track of uploaded emails
            if email:
                uploaded_emails.add(email)

            if is_valid:
                valid_records += 1
            else:
                invalid_records += 1

            error_message = ", ".join(errors)

            records_to_create.append(
                UploadRecord(
                    batch=batch,
                    row_data=row_data,
                    is_valid=is_valid,
                    error_message=error_message
                )
            )

            total_records += 1

            if len(records_to_create) >= batch_size:
                UploadRecord.objects.bulk_create(
                    records_to_create,
                    batch_size=1000
                )
                records_to_create.clear()

        if records_to_create:
            UploadRecord.objects.bulk_create(
                records_to_create,
                batch_size=1000
            )

        batch.total_records = total_records
        batch.valid_records = valid_records
        batch.invalid_records = invalid_records
        batch.status = "completed"

        batch.save(
            update_fields=[
                "total_records",
                "valid_records",
                "invalid_records",
                "status"
            ]
        )

    except Exception as exc:
        batch.status = "failed"
        batch.save(update_fields=["status"])
        print(
            f"Upload processing failed for batch {batch_id}: {exc}"
        )
        import traceback
        traceback.print_exc()

class UploadBatchStatusAPIView(APIView):

    def get(self, request, batch_id):

        try:

            batch = UploadBatch.objects.get(id=batch_id)

            return Response({
                "status": batch.status,
                "total_records": batch.total_records,
                "valid_records": batch.valid_records,
                "invalid_records": batch.invalid_records
        })

        except UploadBatch.DoesNotExist:

            return Response(
                {
                    "error": "Batch not found"
                },
                status=404
            )


class UploadBatchAPIView(APIView):
   

    def post(self, request):
        print("=" * 80)
        print("UploadBatchAPIView POST CALLED")
        print("User:", request.user)
        print("Authenticated:", request.user.is_authenticated)
        print("FILES:", request.FILES)
        print("POST:", request.POST)

        file = request.FILES.get("file")

        if not file:
            return Response({
                "error": "No file uploaded"
            }, status=400)

        batch_name = request.data.get(
            "batch_name",
            ""
        ).strip()

        if not batch_name:
            return Response({
                "error": "Batch name is required"
            }, status=400)

        exhibitor = resolve_exhibitor_from_request(request)

        batch = UploadBatch.objects.create(
            batch_name=batch_name,
            file_name=file.name,
            exhibitor=exhibitor
        )

        file_bytes = file.read()
        columns = []

        is_csv = file.name.lower().endswith('.csv')

        if is_csv:
            try:
                import csv
                from io import StringIO
                try:
                    text = file_bytes.decode('utf-8-sig')
                except UnicodeDecodeError:
                    text = file_bytes.decode('latin-1')
                
                reader = csv.reader(StringIO(text))
                try:
                    columns = next(reader)
                except StopIteration:
                    columns = []
                columns = [str(c).strip() if c is not None else "" for c in columns]
            except Exception as e:
                print("Failed to parse CSV columns:", e)
                columns = []
        else:
            try:
                workbook = load_workbook(
                    BytesIO(file_bytes),
                    read_only=True,
                    data_only=True
                )
                sheet = workbook.active
                columns = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
                workbook.close()
            except Exception:
                columns = []

        thread = threading.Thread(
            target=process_upload_batch,
            args=(batch.id, file.name, file_bytes, columns),
            daemon=True
        )
        thread.start()

        return Response({
            "batch_id": batch.id,
            "batch_name": batch.batch_name,
            "records_created": 0,
            "status": "processing",
            "columns": columns
        })
        
        
from rest_framework.views import APIView
from rest_framework.response import Response

from .models import UploadRecord
from .serializers import UploadRecordSerializer


class ExhibitorUploadBatchesAPIView(APIView):

    def get(self, request):
        exhibitor = resolve_exhibitor_from_request(request)

        if not exhibitor:
            return Response({"batches": []})

        batches = UploadBatch.objects.filter(
            exhibitor=exhibitor
        ).order_by('-uploaded_at', '-id')

        data = []
        for batch in batches:
            data.append({
                "id": batch.id,
                "batch_name": batch.batch_name,
                "file_name": batch.file_name,
                "total_records": batch.total_records,
                "valid_records": batch.valid_records,
                "invalid_records": batch.invalid_records,
                "uploaded_at": batch.uploaded_at.isoformat() if batch.uploaded_at else None
            })

        return Response({"batches": data})

from django.db.models import Q
from rest_framework.response import Response
from rest_framework.views import APIView


class UploadRecordListAPIView(APIView):
    
    def get(self, request):
        print("=" * 80)
        print("UploadRecordListAPIView called")
        print("GET Params:", request.GET)
        
        print("=" * 60)
        print("User:", request.user)
        print("Authenticated:", request.user.is_authenticated)
        print("Session:", request.session.items())
        print("Resolved Exhibitor:", resolve_exhibitor_from_request(request))

        exhibitor = resolve_exhibitor_from_request(request)
        print("Resolved Exhibitor:", exhibitor)

        

        if not exhibitor:
            return Response({
                "draw": 1,
                "count": 0,
                "recordsTotal": 0,
                "recordsFiltered": 0,
                "page": 1,
                "page_size": 100,
                "has_more": False,
                "results": [],
                "data": []
            })

        records = (
            UploadRecord.objects
            .select_related("batch")
            .filter(batch__exhibitor=exhibitor)
            .order_by("-id")
        )

        # ------------------------
        # Keyword Search
        # ------------------------
        from django.db.models import Q

        search = request.GET.get("search", "").strip()

        if search:
            records = records.filter(
                Q(row_data__icontains=search)
            )
            
            
        record_type = request.GET.get("status")

        if record_type == "valid":
            records = records.filter(is_valid=True)

        elif record_type == "invalid":
            records = records.filter(is_valid=False)

        registration_status = request.GET.get("registration_status")

        if registration_status:
            if registration_status == "confirmed":
                records = records.filter(is_valid=True)
            elif registration_status == "pending":
                records = records.filter(is_valid=False)

        # ------------------------
        # Pagination
        # ------------------------
        page_size = int(
            request.GET.get("page_size")
            or request.GET.get("length")
            or 100
        )

        page_size = min(max(page_size, 1), 1000)

        page = int(request.GET.get("page", 1))

        start_param = request.GET.get("start")

        if start_param is not None:
            page = (int(start_param) // page_size) + 1

        total = records.count()

        start = (page - 1) * page_size
        end = start + page_size

        page_records = records[start:end]

        serializer = UploadRecordSerializer(
            page_records,
            many=True
        )

        return Response({
            "draw": int(request.GET.get("draw", 1)),
            "count": total,
            "recordsTotal": total,
            "recordsFiltered": total,
            "page": page,
            "page_size": page_size,
            "has_more": end < total,
            "results": serializer.data,
            "data": serializer.data
        })
        
        
import re

from django.core.exceptions import ValidationError
from django.shortcuts import get_object_or_404

from rest_framework.views import APIView
from rest_framework.response import Response

from .models import UploadRecord
from .serializers import UploadRecordSerializer


class UploadRecordUpdateAPIView(APIView):

    def patch(self, request, record_id):

        record = get_object_or_404(
            UploadRecord,
            id=record_id
        )

        row_data = request.data.get("row_data", {})

        if row_data:
            record.row_data.update(row_data)

        errors = []

        first_name = (record.row_data.get("First Name") or "").strip()
        email = (record.row_data.get("Email") or "").strip()
        company = (record.row_data.get("Company") or "").strip()
        phone = (record.row_data.get("Phone") or "").strip()

        
        if not first_name:
            errors.append("First Name is required")

        
        if not email:
            errors.append("Email is required")

        elif not re.match(
            r"^[^@\s]+@[^@\s]+\.[^@\s]+$",
            email
        ):
            errors.append("Invalid Email")

        elif Visitor.objects.filter(email__iexact=email).exists():
            errors.append("Email already exists")

        
        if not company:
            errors.append("Company is required")

        
        if not phone:
            errors.append("Phone number is required")

        else:
            try:
                phone_validator(phone)
            except ValidationError as e:
                errors.append(e.messages[0])

        record.is_valid = len(errors) == 0
        record.error_message = ", ".join(errors)

        record.save()

        serializer = UploadRecordSerializer(record)

        return Response(serializer.data)
               
class SaveMappingAPIView(APIView):


    def post(self, request, batch_id):
        print("BATCH ID:", batch_id)
        print("DATA:", request.data)

        batch = UploadBatch.objects.get(
            id=batch_id
        )

        mappings = request.data

        ColumnMapping.objects.filter(
            batch=batch
        ).delete()

        for uploaded_column, system_field in mappings.items():

            ColumnMapping.objects.create(
                batch=batch,
                uploaded_column=uploaded_column,
                system_field=system_field
            )

        mapping_dict = {}

        saved_mappings = ColumnMapping.objects.filter(
            batch=batch
        )

        for mapping in saved_mappings:

            mapping_dict[
                mapping.uploaded_column
            ] = mapping.system_field

        records = UploadRecord.objects.filter(
            batch=batch
        )

        records_to_update = []

        for record in records:

            mapped_data = {}

            for key, value in record.row_data.items():

                new_key = mapping_dict.get(
                    key,
                    key
                )

                mapped_data[new_key] = value

            record.row_data = mapped_data

            errors = []

            first_name = mapped_data.get(
                "First Name"
            )

            email = mapped_data.get(
                "Email"
            )

            if not first_name:
                errors.append(
                    "First Name is required"
                )

            if not email:
                errors.append(
                    "Email is required"
                )

            record.is_valid = (
                len(errors) == 0
            )

            record.error_message = (
                ", ".join(errors)
            )

            records_to_update.append(
                record
            )

        UploadRecord.objects.bulk_update(
            records_to_update,
            [
                "row_data",
                "is_valid",
                "error_message"
            ],
            batch_size=1000
        )

        return Response({
            "message": "Mapping saved and applied"
        })

        
        
        
class MappingListAPIView(APIView):

    def get(self, request, batch_id):

        mappings = ColumnMapping.objects.filter(
            batch_id=batch_id
        )

        data = {}

        for mapping in mappings:

            data[
                mapping.uploaded_column
            ] = mapping.system_field

        return Response(data)
    
    
    
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import Invitation, TicketType


class SendInvitationAPIView(APIView):
    

    def post(self, request):
        print("=" * 50)
        print("SEND INVITATION")
        print("User:", request.user)
        print("===== SEND INVITATION VIEW HIT =====")
        
        print("Exhibitor:", getattr(request.user, "exhibitor", None))

        invitations = request.data.get("invitations", [])

        if not invitations:

            return Response(
                {
                    "error": "No invitations provided."
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        success = []
        failed = []

        for data in invitations:

            first_name = data.get("first_name")
            last_name = data.get("last_name")
            email = data.get("email")
            ticket_id = data.get("ticket")

            if not first_name or not last_name or not email or not ticket_id:

                failed.append({
                    "email": email,
                    "error": "Missing required fields."
                })
                continue

            if Invitation.objects.filter(email=email).exists():

                failed.append({
                    "email": email,
                    "error": "Invitation already exists."
                })
                continue

            try:

                ticket = TicketType.objects.get(
                    id=ticket_id
                )

            except TicketType.DoesNotExist:

                failed.append({
                    "email": email,
                    "error": "Invalid ticket."
                })
                continue

            invitation = Invitation.objects.create(

                exhibitor=getattr(request.user, 'exhibitor', None),
                first_name=first_name,

                last_name=last_name,

                email=email,

                ticket=ticket

            )
            print(
                "Saved invitation:",
                invitation.id,
                invitation.exhibitor
            )

            invitation_link = (
                f"http://127.0.0.1:5500/register.html"
                f"?token={invitation.invitation_token}"
            )

            success.append({

                "email": email,

                "invitation_link": invitation_link

            })

        return Response(

            {

                "message": "Invitations processed successfully.",

                "success_count": len(success),

                "failed_count": len(failed),

                "success": success,

                "failed": failed

            },

            status=status.HTTP_201_CREATED

        )
        
from rest_framework.views import APIView
from rest_framework.response import Response

from .models import TicketType

from .serializers import TicketTypeSerializer


class TicketTypeAPIView(APIView):

    def get(self, request):

        tickets = TicketType.objects.all()

        serializer = TicketTypeSerializer(
            tickets,
            many=True
        )

        return Response(serializer.data)
    
    
    
    
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import Badge

from .serializers import BadgeRegistrationSerializer
from rest_framework.permissions import IsAuthenticated

class CreateBadgeAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        exhibitor = request.user.exhibitor

        used = Badge.objects.filter(
            exhibitor=exhibitor
        ).count()

        if used >= exhibitor.allocated_badges:

            return Response(
                {
                    "message": "Badge allocation exceeded."
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = BadgeRegistrationSerializer(
            data=request.data
        )

        if serializer.is_valid():

            serializer.save(
                exhibitor=exhibitor,
                status="confirmed"
            )

            return Response(
                {
                    "message": "Badge created successfully.",
                    "data": serializer.data
                },
                status=status.HTTP_201_CREATED
            )

        return Response(
            serializer.errors,
            status=status.HTTP_400_BAD_REQUEST
        )
        
        
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .serializers import CreateExhibitorSerializer


class CreateExhibitorAPIView(APIView):

    def post(self, request):

        serializer = CreateExhibitorSerializer(
            data=request.data
        )

        if serializer.is_valid():

            exhibitor = serializer.save()

            return Response(
                {
                    "message": "Exhibitor created successfully.",
                    "company": exhibitor.company_name,
                    "allocated_badges": exhibitor.allocated_badges
                },
                status=status.HTTP_201_CREATED
            )

        return Response(
            serializer.errors,
            status=status.HTTP_400_BAD_REQUEST
        )
        

import json
from django.contrib.auth import authenticate, login
from django.http import JsonResponse
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_protect


@method_decorator(csrf_protect, name="dispatch")
class LoginAPIView(View):
    http_method_names = ["post"]

    def post(self, request, *args, **kwargs):
        if request.content_type == "application/json":
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse(
                    {"success": False, "errors": {"__all__": "Invalid JSON body."}},
                    status=400,
                )
        else:
            data = request.POST

        username = (data.get("username") or "").strip()
        password = data.get("password") or ""

        errors = {}

        if not username:
            errors["username"] = "Username is required."
        elif len(username) > 30:
            errors["username"] = "Username must be at most 30 characters."

        if not password:
            errors["password"] = "Password is required."
        elif len(password) < 8:
            errors["password"] = "Password must be at least 8 characters."
        elif len(password) > 128:
            errors["password"] = "Password must be at most 128 characters."

        if errors:
            return JsonResponse({"success": False, "errors": errors}, status=400)

        user = authenticate(request, username=username, password=password)

        if user is None:
            return JsonResponse(
                {"success": False, "errors": {"__all__": "Invalid username or password."}},
                status=401,
            )

        if not user.is_active:
            return JsonResponse(
                {"success": False, "errors": {"__all__": "This account is inactive."}},
                status=403,
            )

        login(request, user) 

        return JsonResponse({
            "success": True,
            "user": {
                "id": user.id,
                "username": user.username,
                "is_staff": user.is_staff,
            },
        })

def exhibitor_dashboard_view(request):
    return render(request, 'exhibitor.html')


def exhibitor_login_page_view(request):
    return render(request, 'exhibitor_login.html')


def visitor(request):
    return render(request, 'visitor.html')




from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import UploadBatch
from .serializers import UploadBatchSerializer


class UploadBatchListAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        exhibitor = request.user.exhibitor

        batches = UploadBatch.objects.filter(
            exhibitor=exhibitor
        ).order_by("-uploaded_at")

        serializer = UploadBatchSerializer(
            batches,
            many=True
        )

        return Response(serializer.data)    
    
    
    
    
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import Exhibitor
from .serializers import AllocateBadgeSerializer


class AllocateBadgeAPIView(APIView):

    def patch(self, request, exhibitor_id):

        try:
            exhibitor = Exhibitor.objects.get(
                id=exhibitor_id
            )

        except Exhibitor.DoesNotExist:

            return Response(
                {
                    "message": "Exhibitor not found."
                },
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = AllocateBadgeSerializer(
            exhibitor,
            data=request.data,
            partial=True
        )

        if serializer.is_valid():

            serializer.save()

            return Response(
                {
                    "message": "Badge allocation updated successfully.",
                    "exhibitor": {
                        "id": exhibitor.id,
                        "company_name": exhibitor.company_name,
                        "contact_person": exhibitor.contact_person,
                        "allocated_badges": serializer.data["allocated_badges"]
                    }
                }
            )

        return Response(
            serializer.errors,
            status=status.HTTP_400_BAD_REQUEST
        )
        
        
        
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import Invitation, Badge
from .serializers import InvitationSerializer


class CreateInvitationAPIView(APIView):

    def post(self, request):

        if Badge.objects.filter(
            email=request.data.get("email")
        ).exists():

            return Response(
                {
                    "message": "Badge already exists for this email."
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        if Invitation.objects.filter(
            email=request.data.get("email")
        ).exists():

            return Response(
                {
                    "message": "Invitation already sent."
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = InvitationSerializer(
            data=request.data
        )

        if serializer.is_valid():

            invitation = serializer.save(
                exhibitor=request.user.exhibitor
            )

            invitation_link = request.build_absolute_uri(
                f"/register/{invitation.invitation_token}/"
            )

            return Response(
                {
                    "message": "Invitation created successfully.",
                    "data": serializer.data,
                    "invitation_link": invitation_link
                },
                status=status.HTTP_201_CREATED
            )

        return Response(
            serializer.errors,
            status=status.HTTP_400_BAD_REQUEST
        )
        
        
        
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import Invitation
from .serializers import InvitationListSerializer


class InvitationListAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):
        print("=" * 50)
        print("LIST INVITATIONS")

        exhibitor = request.user.exhibitor

        print("User:", request.user)
        print("Exhibitor:", exhibitor)

        invitations = Invitation.objects.filter(
            exhibitor=exhibitor
        )

        print("Count:", invitations.count())

        exhibitor = getattr(request.user, 'exhibitor', None)

        if not exhibitor:
            return Response([], status=status.HTTP_200_OK)

        invitations = Invitation.objects.filter(
            exhibitor=exhibitor
        ).order_by("-created_at")

        serializer = InvitationListSerializer(
            invitations,
            many=True,
            context={
                "request": request
            }
        )

        return Response(serializer.data)
    
    
    
def visitor(request):

    token = request.GET.get("token")

    invitation = Invitation.objects.filter(
        invitation_token=token,
        status="Invited",
        is_used=False
    ).first()

    return render(
        request,
        "visitor.html",
        {
            "invitation": invitation
        }
    )
    
    

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.utils import timezone
from django.core.validators import RegexValidator
from django.core.exceptions import ValidationError

from .models import Invitation, Visitor


phone_validator = RegexValidator(
    regex=r"^\+?[0-9\s\-]{7,15}$",
    message="Enter a valid phone number."
)


class GetInvitationAPIView(APIView):
    """GET /api/invitation/<uuid:token>/ -> used by register.html to prefill data"""

    def get(self, request, token):
        try:
            invitation = Invitation.objects.select_related(
                "ticket", "exhibitor"
            ).get(invitation_token=token)
        except Invitation.DoesNotExist:
            return Response(
                {"error": "Invalid invitation link."},
                status=status.HTTP_404_NOT_FOUND
            )

        if invitation.expires_at and invitation.expires_at < timezone.now():
            return Response(
                {"error": "This invitation link has expired."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if invitation.is_used or invitation.registered_count >= invitation.link_limit:
            return Response(
                {"error": "This invitation has already been used."},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response({
            "first_name": invitation.first_name,
            "last_name": invitation.last_name,
            "email": invitation.email,
            "ticket_name": invitation.ticket.name if hasattr(invitation.ticket, "name") else str(invitation.ticket),
            "exhibitor_name": str(invitation.exhibitor) if invitation.exhibitor else None,
            "status": invitation.status,
        }, status=status.HTTP_200_OK)


class CompleteRegistrationAPIView(APIView):
    """POST /api/register/<uuid:token>/ -> creates the Visitor record"""

    def post(self, request, token):
        print("Inside CompleteRegistrationAPIView")
        try:
            invitation = Invitation.objects.select_related(
                "ticket", "exhibitor"
            ).get(invitation_token=token)
        except Invitation.DoesNotExist:
            return Response(
                {"error": "Invalid invitation link."},
                status=status.HTTP_404_NOT_FOUND
            )

        if invitation.expires_at and invitation.expires_at < timezone.now():
            return Response(
                {"error": "This invitation link has expired."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if invitation.is_used or invitation.registered_count >= invitation.link_limit:
            return Response(
                {"error": "This invitation has already been used."},
                status=status.HTTP_400_BAD_REQUEST
            )

        data = request.data

        company_name = (data.get("company_name") or "").strip()
        job_title = (data.get("job_title") or "").strip()
        phone_number = (data.get("phone_number") or "").strip()
        country_of_residence = (data.get("country_of_residence") or "").strip()
        nationality = (data.get("nationality") or "").strip()

        errors = {}

        if not company_name:
            errors["company_name"] = "Company name is required."
        elif len(company_name) > 255:
            errors["company_name"] = "Company name is too long."

        if not job_title:
            errors["job_title"] = "Job title is required."
        elif len(job_title) > 255:
            errors["job_title"] = "Job title is too long."

        if not phone_number:
            errors["phone_number"] = "Phone number is required."
        else:
            try:
                phone_validator(phone_number)
            except ValidationError as e:
                errors["phone_number"] = e.messages[0]

        if not country_of_residence:
            errors["country_of_residence"] = "Country of residence is required."
        elif len(country_of_residence) > 100:
            errors["country_of_residence"] = "Too long."

        if not nationality:
            errors["nationality"] = "Nationality is required."
        elif len(nationality) > 100:
            errors["nationality"] = "Too long."

        if errors:
            return Response({"errors": errors}, status=status.HTTP_400_BAD_REQUEST)

        if Visitor.objects.filter(email=invitation.email).exists():
            return Response(
                {"error": "A visitor with this email is already registered."},
                status=status.HTTP_400_BAD_REQUEST
            )

        visitor = Visitor.objects.create(
            invitation=invitation,
            exhibitor=invitation.exhibitor,
            ticket=invitation.ticket,
            first_name=invitation.first_name,
            last_name=invitation.last_name,
            email=invitation.email,
            company_name=company_name,
            job_title=job_title,
            phone_number=phone_number,
            country_of_residence=country_of_residence,
            nationality=nationality,
            status="Confirmed",
        )

        invitation.status = "Confirmed"
        invitation.is_used = True
        invitation.registered_count += 1
        invitation.save()

        return Response({
            "message": "Registration successful.",
            "visitor": {
                "id": visitor.id,
                "first_name": visitor.first_name,
                "last_name": visitor.last_name,
                "badge_id": str(visitor.badge_id),
                "status": visitor.status,
            }
        }, status=status.HTTP_201_CREATED)
        
        
def register_view(request):
    return render(request, "register.html")